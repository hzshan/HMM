import pickle
import numpy as np
import os
import time
from utils import load_series
from utils import ProcessedData
from openpyxl import Workbook, load_workbook

"""Check the raw xlsx files for these parameters"""
time_column = 'B'
X_column = 'C'
Y_column = 'D'
velocity_column = 'I'
ConditionID_cell = 'B33'
ratID_cell = 'B34'
Date_cell = 'B27'
opener_cell = 'B35'
open_time_cell = 'B36'
group_cell = 'B39'

"""What's the file name of the raw data?"""
file_header = "Raw data-propranolol Belinda DEF-Trial   "
pickle_name = 'propranolol_data'

"""Tracking details"""
time_step_of_tracking = 0.066
arena_diameter = 60

# %%
loading_start = 1
loading_end = 288

start_time = time.time()
for z in range(loading_start, loading_end + 1):

    loading_video_number = z
    if z < 10:
        name = file_header + '  %s.xlsx' % loading_video_number
    elif 9 < z < 100:
        name = file_header + ' %s.xlsx' % loading_video_number
    else:
        name = file_header + '%s.xlsx' % loading_video_number

    if os.path.isfile(name) is False:
        print('File with name:' + name + 'IS NOT FOUND.')
        continue

    print("Currently processing: " + name)
    workbook = load_workbook(name)
    sheet = workbook.worksheets[0]
    if z == loading_start:
        first_row = int(sheet['B1'].value) + 1

    last_row = sheet.max_row

    condition = sheet[ConditionID_cell].value
    rat_id = sheet[ratID_cell].value
    date = sheet[Date_cell].value

    opener = sheet[opener_cell].value
    opening_time = sheet[open_time_cell].value
    group = sheet[group_cell].value

    time_line = load_series(first_row, last_row, time_column, sheet)
    x_coors = load_series(first_row, last_row, X_column, sheet)
    y_coors = load_series(first_row, last_row, Y_column, sheet)

    save_object = ProcessedData(filename=name,
                                condition=condition,
                                opener=opener,
                                opening_time=opening_time,
                                group=group,
                                rat_id=rat_id,
                                date=date,
                                timeline=time_line,
                                x_coors=x_coors,
                                y_coors=y_coors,
                                timestep=time_step_of_tracking,
                                arena_size=arena_diameter)

    save_name = pickle_name + "%s.pkl" % loading_video_number
    pickle.dump(save_object, open(save_name, 'wb'))

    print("Time elapsed: %05s sec" % (time.time() - start_time))

print('--------------------------------------------------------------')
print('File Processing Complete.')

